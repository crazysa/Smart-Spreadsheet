from helper_functions import get_sheet_from_excel, process_hierarchical_table, get_excel_tables


# This code is provided under the following terms:
# This code is for testing, non-commercial use only.
# This code or any part of it shouldn't be copied without my consent
path_to_excel = "/home/shub/Desktop/Smart-Spreadsheet/tests/example_0.xlsx"
sheet_name = "Analysis Output"

worksheet = get_sheet_from_excel(path_to_excel, sheet_name)


import openpyxl
import csv
import os

def load_excel(file_path):
    return openpyxl.load_workbook(file_path, data_only=True)

def find_tables(sheet):
    tables = []
    rows = list(sheet.iter_rows())

    in_table = False
    table_start = None
    for i, row in enumerate(rows):
        start_table = False
        
        for r in row:
            if r.fill.start_color.index == 4 :
                start_table = True
                break
        if start_table :
            if table_start is not None :
                if table_start != i-1:
                    tables.append((table_start, i - 1))
                    table_start = i
            else:
                table_start = i
    tables.append((table_start, i-1))

    return tables

def extract_table(sheet, start_row, end_row):
    table_data = []
    for row in sheet.iter_rows(min_row=start_row + 1, max_row=end_row + 1, values_only=True):
        table_data.append(list(row))
    return table_data

def save_table_as_csv(table_data, table_name, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    file_path = os.path.join(output_dir, f"{table_name}.csv")
    with open(file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        for row in table_data:
            writer.writerow(row)

def main(excel_file_path, output_dir):
    wb = load_excel(excel_file_path)

    for sheet in wb.worksheets:
        tables = find_tables(sheet)
        for i, (start_row, end_row) in enumerate(tables):
            table_data = extract_table(sheet, start_row, end_row)
            table_name = f"{sheet.title}_table_{i+1}"
            start_index = 0
            end_index = len(table_data[0])
            table_data = [f for f in table_data if any(f)]
            
            for index in range(len(table_data[0])):
                for value in table_data:
                    if value[index] is not None:
                        start_index = index
                        break
                if start_index != 0:
                    break
            for index in range(len(table_data[0])-1, -1, -1):
                for value in table_data:
                    if value[index] is not None:
                        end_index = index
                        break
                if end_index != len(table_data[0]):
                    break
                      
            for index in range(len(table_data)):
                table_data[index] = table_data[index][start_index:end_index]

            save_table_as_csv(table_data, table_name, output_dir)
            print(f"Table {table_name} saved to {output_dir}")


if __name__ == "__main__":

    excel_file_path = "/home/shub/Desktop/Smart-Spreadsheet/tests/example_0.xlsx"
    output_dir = "output_tables"
    main(excel_file_path, output_dir)


