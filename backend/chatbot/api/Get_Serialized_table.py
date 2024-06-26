# from helper_functions import get_sheet_from_excel


# This code is provided under the following terms:
# This code is for testing, non-commercial use only.
# This code or any part of it shouldn't be copied without my consent


# worksheet = get_sheet_from_excel(path_to_excel, sheet_name)

from rest_framework import status

from rest_framework.response import Response
import openpyxl
import csv
import sys
import shutil
import os
from rest_framework.decorators import api_view, renderer_classes
from rest_framework.renderers import JSONRenderer, TemplateHTMLRenderer
from django.core.files.storage import default_storage

def load_excel(file_path):
    return openpyxl.load_workbook(file_path, data_only=True)

def get_remaining_table_from_start(sheet, start_row, start_col):
    row_to_check = list(sheet.iter_rows())[start_row][start_col:]
    end_col = len(row_to_check)-1
    for index,col in enumerate(row_to_check):
        if not col.border.top.style:
            end_col = start_col+index-1
            break
    
    col_to_check = list(sheet.iter_cols())[start_col][start_row:]
    end_row = len(col_to_check)-1
    for index, row in enumerate(col_to_check):
        if not row.border.left.style:
            end_row = start_row+index-1
            break
    return start_row, start_col, end_row, end_col
def find_tables_advanced(sheet):
    tables = []
    
    rows = list(sheet.iter_rows())
    end_col = 0
    end_row = 0
    for i, row in enumerate(rows):
        
        for col_index,r in enumerate(row):
            if r.border.left.style and r.border.top.style  and( col_index >= end_col or i >= end_row):
                start_row, start_col, end_row, end_col = get_remaining_table_from_start(sheet, i, col_index)
                tables.append((start_row, start_col, end_row, end_col))
    

    return tables
    

def find_tables(sheet):
    tables = []
    rows = list(sheet.iter_rows())
    table_start = None
    for i, row in enumerate(rows):
        start_table = False
        
        for r in row:
            if r.fill.start_color.index == 4 :
                start_table = True
                break
        if start_table :
            if table_start is not None :
                if table_start != i-1:
                    tables.append((table_start, i - 1))
                    table_start = i
            else:
                table_start = i
    tables.append((table_start, i-1))

    return tables

def extract_table(sheet, start_row, end_row, start_col=None, end_col=None):
    # If start_col and end_col are not provided, default to the entire row width
    if start_col is None:
        start_col = 1
    if end_col is None:
        end_col = sheet.max_column
    
    table_data = []
    for row in sheet.iter_rows(min_row=start_row + 1, max_row=end_row + 1, min_col=start_col, max_col=end_col, values_only=True):
        table_data.append(list(row))
    return table_data

def save_table_as_csv(table_data, table_name, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    file_path = os.path.join(output_dir, f"{table_name}.csv")
    with open(file_path, mode='w', newline='') as file:
        writer = csv.writer(file)
        for row in table_data:
            writer.writerow(row)

def get_start_index(table_data):
    for index in range(len(table_data[0])):
        start_index = 0
        for value in table_data:
            if value[index] is not None:
                start_index = index
                break
        if start_index != 0:
            break
    return start_index

def get_end_index(table_data):
    end_index = len(table_data[0])
    for index in range(len(table_data[0])-1, -1, -1):
        for value in table_data:
            if value[index] is not None:
                end_index = index
                break
        if end_index != len(table_data[0]):
            break
    return end_index

import random, string
def random_string(num_chars=24):
    return ''.join(random.choice(
        string.ascii_uppercase +
        string.ascii_lowercase + string.digits) for _ in range(num_chars))

def main(excel_file_path, output_dir):
    wb = load_excel(excel_file_path)
    use_the_updated_algo = False
    for sheet in wb.worksheets:
        if not use_the_updated_algo:
            tables = find_tables(sheet)
            for i, (start_row, end_row) in enumerate(tables):
                table_data = extract_table(sheet, start_row, end_row)
                table_name = f"{sheet.title}_table_{i+1}"
                start_index = 0
                end_index = len(table_data[0])
                table_data = [f for f in table_data if any(f)]
                start_index = get_start_index(table_data)
                end_index = get_end_index(table_data)
                            
                for index in range(len(table_data)):
                    table_data[index] = table_data[index][start_index:end_index+1]

                save_table_as_csv(table_data, table_name, output_dir)
                print(f"Table {table_name} saved to {output_dir}")
        else:
            tables = find_tables_advanced(sheet)
            for i, (start_row, start_col, end_row, end_col) in enumerate(tables):
                table_data = extract_table(sheet, start_row, end_row, start_col, end_col)
                table_name = f"{sheet.title}_table_{i+1}"
                start_index = 0
                end_index = len(table_data[0])
                table_data = [f for f in table_data if any(f)]
                # start_index = get_start_index(table_data)
                # end_index = get_end_index(table_data)
                # start_index = 0
            
                for index in range(len(table_data)):
                    table_data[index] = table_data[index][start_index:end_index+1]

                save_table_as_csv(table_data, table_name, output_dir)
                print(f"Table {table_name} saved to {output_dir}")
            



@api_view(['POST'])
@renderer_classes((TemplateHTMLRenderer, JSONRenderer))
def upload_excel_file (request):
    print(request)
    excel_path = f"{os.getcwd()}/tmp/{random_string()}"    
    default_storage.save(os.path.join(excel_path, "excel.xlsx"), request.FILES.get('files'))
    output_dir  = f"{os.getcwd()}/tmp/output/"
    if os.path.isdir(output_dir):
        shutil.rmtree(output_dir)
        os.makedirs(output_dir)
    main(os.path.join(excel_path, "excel.xlsx"), output_dir)
    # pass
    return Response(
            status=status.HTTP_200_OK,
            data="Testing, Please check your input data and make sure the frames are of good quality before trying again")
    


